#CSV demo
#https://github.com/Netuser-Py/pythoncsvdemo
#CSV is native but has differnt flavors
# Watch out for:
#    UNIX vs Windows vs Apple
#    comma or tab delimited
# Note: The reader is hard-coded to recognise either '\r' or '\n' as end-of-line, 
#       and ignores lineterminator. This behavior may change in the future. 
# Be ready to do pre processing to get your file "just Right"
# How will you handle ""?
#

import csv
import tempfile
import openpyxl

Hold_recs = {} #hold space/dict for input data
    
GWS_target = tempfile.TemporaryFile()
#print (csv.list_dialects())

def load_CLNT():
#    pre-process: replace single b'\n' with b'|' to eleiminate lonely \n in multi-row cells
#    uses tempfile.TemporaryFile() 
    WS_CLNT_file = r'CLNTCodesDomo1.csv'
    fp = open(WS_CLNT_file,'rb')
    file_txt = fp.read()
    file_txt = file_txt.replace(b'\r',b'^').replace(b'\n',b'|').replace(b'^|',b'\r\n')
    GWS_target.write(file_txt)
    
def write_test():
# write tempfile to disk
    WS_CLNT_file_out = r"CLNTCodesDomo1_out.csv"
    GWS_target.seek(0)    
    file_txt1 = GWS_target.read()
    fp1 = open(WS_CLNT_file_out,'wb')
    fp1.write(file_txt1)

def input_CSVfile():   
    with open(r'CLNTCodesDomo1_out.csv','r',newline='') as csvfile:
        dialect = csv.Sniffer().sniff(csvfile.read(1024)) #use niffer to identify the dialect
        csvfile.seek(0) #go to the top
        print (dialect)
        reader = csv.reader(csvfile, dialect, lineterminator='\r\n') #ignores lineterminator
        # ... process CSV file contents here ...
        i=0
        for row in reader:
            i += 1
            Hold_recs.update({i:row})

def Make_sheet():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
# add text data from hardcoded data in program
    ws.title = "Test Sheet"
    ws['A1']= "1-1"
    ws['A2']= "1-2"
    ws['A3']= "1-3"
    ws['A4']= "1-4"
    ws['A5']= "1-5"
    ws['B1']= "2-1"
    ws['B2']= "2-2"
    ws['B3']= "2-3"
    ws['B4']= "2-4"
    ws['B5']= "2-5"
    
# Add rows from Hold_recs
    for row in Hold_recs:
         ws.append(Hold_recs[row])
    
    wb.save(r'test.xlsx')
    
if __name__  ==  '__main__':
   load_CLNT()
   write_test()
   input_CSVfile()
   Make_sheet()
   print('\a')
